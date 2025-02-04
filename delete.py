import openpyxl
import os


class ClearData:
    def __init__(self, file_name="./storage/new_library.xlsx"):
        self.file_name = file_name
        self.workbook = None
        self.worksheet = None
        self.load_workbook()

    def load_workbook(self):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.worksheet = self.workbook.active  
            

    def delete_row(self, row_number):
        """Deletes the entire row."""
        self.worksheet.delete_rows(row_number)
        self.workbook.save(self.file_name)

    def delete_cell(self, cell_name):
        """Delete the content of a specific cell."""
        self.worksheet[cell_name].value = None  
        self.workbook.save(self.file_name)

    def deledata(self):
        user_input = input("Enter the row number or cell (e.g., '3' or 'A2'): ")
        if user_input.isdigit():  
            self.delete_row(int(user_input))  
        else:  
            self.delete_cell(user_input) 

def main():
    manager = ClearData()
    manager.deledata()

if __name__ == "__main__":
    main()
