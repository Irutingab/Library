import openpyxl
import os

class LibraryManager:
    def __init__(self, file_name="./storage/new_library.xlsx", sheet_name="Information_Librarie_SL"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.load_workbook()

    def load_workbook(self):
        """Loads the workbook or creates a new one if it does not exist."""
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            if self.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.sheet_name]
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = self.sheet_name
            self.worksheet.append(["Title", "Author", "Publication Year"])
            self.save_workbook()

    def find_book_row(self, book_title):
        """Finds the row number of a book based on the title without loading the entire sheet."""
        for row_num, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and row[0].strip() == book_title:
                return row_num
        return None  

    def save_workbook(self):
        """Saves the workbook."""
        self.workbook.save(self.file_name)

    def add_book(self):
        """Adds a new book to the excel file."""
        book = input("Enter book title: ").strip()
        author = input("Enter author name: ").strip()
        publication_year = input("Enter publication year: ").strip()

        if book and author and publication_year:
            self.worksheet.append([book, author, publication_year])
            print(f"Book '{book}' added successfully!")
        else:
            print("Error: All fields must be completed!")

    def view_books(self):
        """Displays books in chunks to reduce memory usage."""
        print("Library Collection:\n")
        chunk_size = 50  
        row_start =  2 
        total_rows = self.worksheet.max_row

        while True:
            rows = list(self.worksheet.iter_rows(min_row=row_start, max_row=row_start + chunk_size, values_only=True))

            if not rows:
                print("No more books to display.")
                break

            for i, row in enumerate(rows, start=row_start):
                print(f"{i}: {row[0]}, Author: {row[1]}, Year: {row[2]}")

            row_start += chunk_size

            user_input = input("\nPress Enter to see more or 'q' to quit: ").strip().lower()
            if user_input == 'q':
                break

    def edit_book(self):
        """Allows a user to edit book details."""
        book_to_edit = input("Enter the title of the book you want to edit: ").strip()

        row_num = self.find_book_row(book_to_edit)
        if row_num:
            print("What would you like to edit?")
            print("1. Title")
            print("2. Author")
            print("3. Publication Year")
            print("4. Cancel")

            choice = input("Enter your choice (1-4): ").strip()

            if choice == '1':
                new_title = input("Enter new book title: ").strip()
                self.worksheet.cell(row=row_num, column=1).value = new_title
                print("Title updated successfully!")

            elif choice == '2':
                new_author = input("Enter new author name: ").strip()
                self.worksheet.cell(row=row_num, column=2).value = new_author
                print("Author updated successfully!")

            elif choice == '3':
                new_year = input("Enter new publication year: ").strip()
                self.worksheet.cell(row=row_num, column=3).value = new_year
                print("Publication year updated successfully!")

            elif choice == '4':
                print("Edit cancelled.")
                return

            self.save_workbook()
            print(f"Book '{book_to_edit}' has been updated successfully!")
        else:
            print("Book not found.")

    def delete_book(self):
        """Deletes a book from the library."""
        book_to_delete = input("Enter the title of the book you want to delete: ").strip()

        row_num = self.find_book_row(book_to_delete)
        if row_num:
            self.worksheet.delete_rows(row_num)
            self.save_workbook()
            print(f"Book '{book_to_delete}' deleted successfully!")
        else:
            print("Book not found.")

    def display_menu(self):
        """Display user choices."""
        while True:
            print("\nLibrary Management Menu:")
            print("1. Add New Book")
            print("2. Edit Existing Book")
            print("3. Delete Existing Book")
            print("4. View All Books")
            print("5. Save and Exit")

            choice = input("Enter your choice (1-5): ")

            if choice == '1':
                self.add_book()
            elif choice == '2':
                self.edit_book()
            elif choice == '3':
                self.delete_book()
            elif choice == '4':
                self.view_books()
            elif choice == '5':
                print("Saving changes and exiting...")
                self.save_workbook()
                break
            else:
                print("Invalid choice. Please try again.")

def main():
    manager = LibraryManager()
    manager.display_menu()

if __name__ == "__main__":
    main()
