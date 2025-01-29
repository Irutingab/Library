from faker import Faker
import openpyxl
import os

class LibraryManagement:
    def __init__(self, file_name="./storage/Librarie_Saint_Laurent.xlsx", sheet_name="Information_Librarie_SL"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.faker = Faker()
        self.__initialize_file()

    def __initialize_file(self):
        """Create the file if it does not exist, or load it if it does."""
        if not os.path.exists(self.file_name):
            self.workbook = openpyxl.Workbook()#creates a new, and empty workbook
            self.worksheet = self.workbook.active
            self.worksheet.title = self.sheet_name
            self.worksheet.append(["Book_name", "Author_name", "Publication_Year"])
            self.workbook.save(self.file_name)
            print(f"File '{self.file_name}' created with sheet '{self.sheet_name}' and headers.")
        else:
            self.workbook = openpyxl.load_workbook(self.file_name)#loads an existing workbook
            if self.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.sheet_name]
            else:
                self.worksheet = self.workbook.create_sheet(title=self.sheet_name)
                self.worksheet.append(["Book_name", "Author_name", "Publication_Year"])
                print(f"Sheet '{self.sheet_name}' added to '{self.file_name}'.")

    def generate_data(self):
        """Generate book info using Faker."""
        Book_name = self.faker.sentence(nb_words=4)
        Author_name = self.faker.name()
        Publication_Year = self.faker.year()
        return Book_name, Author_name, Publication_Year

    def save_data_to_sheet(self):
        """Generate and save book data to the sheet."""
        Book_name, Author_name, Publication_Year = self.generate_data()
        self.worksheet.append([Book_name, Author_name, Publication_Year])
        self.workbook.save(self.file_name)
        print(f"Data saved to '{self.file_name}':")
        print(f"Book Name: {Book_name}")
        print(f"Author Name: {Author_name}")
        print(f"Publication Year: {Publication_Year}")
        
        

