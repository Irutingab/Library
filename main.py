import openpyxl
import os

class LibraryManager:
    def __init__(self, file_name="./storage/new_library.xlsx", sheet_name="Information_Librarie_SL"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None  
        self.load_workbook()

    def load_workbook(self):
        """Loads the workbook or creates a new one if it does not exist."""
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            if self.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = self.sheet_name
            self.worksheet.append(["Title", "Author", "Publication Year"])  
            self.save_workbook()

    def save_workbook(self):
        """Saves the workbook to the file."""
        self.workbook.save(self.file_name)  

    def add_book(self):
        """Adds a new book to the excel file."""
        book = input("Enter book title: ").strip()
        author = input("Enter author name: ").strip()
        publication_year = input("Enter publication year: ").strip()

        if book and author and publication_year:
            self.worksheet.append([book, author, publication_year])  
            print(f"Book '{book}' added successfully!")
        else:
            print("Error: All fields must be completed!")

    def view_books(self):
        """Displays all books in the library."""
        print("Library Collection:")
        for row in self.worksheet.iter_rows(values_only=True):
            print(row)

    def edit_book(self):
        """Allows a user to edit specified details of an existing book."""
        book_to_edit = input("Enter the title of the book you want to edit: ").strip()
        found = False

        for row in self.worksheet.iter_rows(min_row=2, values_only=False):  
            if row[0].value == book_to_edit:
                found = True
                print("What would you like to edit?")
                print("1. Title")
                print("2. Author")
                print("3. Publication Year")
                print("4. Cancel")

                choice = input("Enter your choice (1-4): ").strip()

                if choice == '1':
                    new_title = input("Enter new book title: ").strip()
                
                    row[0].value = new_title
                    print("Title updated successfully!")

                elif choice == '2':
                    new_author = input("Enter new author name: ").strip()
                    row[1].value = new_author
                    print("Author updated successfully!")

                elif choice == '3':
                    new_year = input("Enter new publication year: ").strip()
                    row[2].value = new_year
                    print("Publication year updated successfully!")

                elif choice == '4':
                    print("Edit cancelled.")
                    return

                else:
                    print("Invalid choice. Returning to menu.")
                    return

                self.save_workbook()
                print(f"Book '{book_to_edit}' has been updated successfully!")
                return  

        if not found:
            print("Book not found.")

    def display_menu(self):
        """Display user choices."""
        while True:
            print("\nLibrary Management Menu:")
            print("1. Add New Book")
            print("2. Edit Existing Book")
            print("3. View All Books")  
            print("4. Save and Exit")

            choice = input("Enter your choice (1-4): ")

            if choice == '1':
                self.add_book()
            elif choice == '2':
                self.edit_book()
            elif choice == '3':
                self.view_books()
            elif choice == '4':
                print("Saving changes and exiting...")
                self.save_workbook()
                break
            else:
                print("Invalid choice. Please try again.")

def main():
    manager = LibraryManager()
    manager.display_menu()

if __name__ == "__main__":
    main()
