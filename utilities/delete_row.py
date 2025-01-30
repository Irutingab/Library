import openpyxl
import os

class DeleteRowFromFile:
    def __init__(self, file_name = "./storage/Librarie_Saint_Laurent.xlsx"):
        self.file_name = file_name
        self.workbook = None
        self.worksheet = None
        
    def RemoveRow(self,):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.worksheet = self.workbook.active
            
            #Delete a specified rows
            self.worksheet.delete_rows(4, amount=6)
            self.workbook.save(self.file_name)
            print(f"row deleted from '{self.file_name}'")  
            
    def RemoveCol(self,):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.worksheet = self.workbook.active
            
            self.worksheet.delete_cols(3) 
            self.workbook.save(self.file_name)
            print(f"column deleted from '{self.file_name}'") 
            
            
    def RemovecellValue(self,):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.worksheet = self.workbook.active
            
            
            self.worksheet["B3"].value = None #deletes the value of cell B3
            self.workbook.save(self.file_name)
            print(f"cell deleted from '{self.file_name}'")  
            
            
def main():
        manager = DeleteRowFromFile()
        #manager.RemovecellValue()
        #manager.RemoveRow()
        #manager.RemoveCol()
if __name__ == "__main__":
        main()      