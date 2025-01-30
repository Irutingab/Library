import openpyxl
import os
from openpyxl import workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class CountBooks:
    def __init__(self, file_name="./storage/Librarie_Saint_Laurent.xlsx"):
        self.file_name = file_name
        self.workbook = None
        self.worksheet = None

    def AvailableBooks(self):
        if os.path.exists(self.file_name):
            # Load the workbook
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.worksheet = self.workbook.active
            
            # Get the maximum row number
            max_row = self.worksheet.max_row
            
            # Add header and formula
            self.worksheet["E1"] = "Number of available books"
            self.worksheet["E2"] = f'=COUNTA(A2:A{max_row})'  
            
            # Save changes
            self.workbook.save(self.file_name)
            print(f"Books counted and successfully saved in '{self.file_name}'")
        else:
            print(f"Error: File '{self.file_name}' not found")