from faker import Faker
import openpyxl
import os
from openpyxl.utils import get_column_letter, column_index_from_string

class RecoverColumns:
    def __init__(self, column_name="C", column_index="C", file_name="./storage/Librarie_Saint_Laurent.xlsx", sheet_name="Information_Librarie_SL"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.column_name = column_name
        self.column_index = column_index
        self.workbook = None
        self.worksheet = None
        self.faker = Faker()
        
    def NewrowsAndcolumns(self):
        """Add data in D"""
        if os.path.exists(self.file_name):
            # Load existing workbook instead of creating new one
            self.workbook = openpyxl.load_workbook(self.file_name)
            # Get the worksheet
            self.worksheet = self.workbook[self.sheet_name]
            
            # Convert column name to index if column_index is a letter
            if isinstance(self.column_index, str):
                self.column_index = column_index_from_string(self.column_index)
            
            # Add header
            self.worksheet.cell(row=1, column=self.column_index, value="Publication_Year")
            
            # Add data to all rows
            for row in range(2, self.worksheet.max_row + 1):
                self.worksheet.cell(row=row, column=self.column_index, value=self.faker.year())
            
            # Save the workbook
            self.workbook.save(self.file_name)
            print(f"Column '{self.column_name}' updated successfully in '{self.file_name}'.")
        else:
            print(f"File '{self.file_name}' does not exist.")
            
