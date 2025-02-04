import openpyxl
import os

class LibraryManager:
    def __init__(self, file_name="./storage/new_library.xlsx", sheet_name="Information_Librarie_SL"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None  
        self.load_workbook()

    def load_workbook(self):
        """Loads the workbook or creates a new one if it does not exist."""
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            if self.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = self.sheet_name
            self.worksheet.append(["Title", "Author", "Publication Year"])  
            self.save_workbook()

    def save_workbook(self):
        """Saves the workbook to the file."""
        self.workbook.save(self.file_name)  

    def add_book(self):
        """Adds a new book to the excel file."""
        book = input("Enter book title: ").strip()
        author = input("Enter author name: ").strip()
        publication_year = input("Enter publication year: ").strip()

        if book and author and publication_year:
            self.worksheet.append([book, author, publication_year])  
            print(f"Book '{book}' added successfully!")
        else:
            print("Error: All fields must be Completed!")

    def view_books(self):
        """Displays all books in the library."""
        print("Library Collection:")
        for row in self.worksheet.iter_rows(values_only=True):
            print(row)

    def edit_book(self):
        """Allows a user to edit an existing book by title."""
        book_to_edit = input("Enter the title of the book you want to edit: ").strip()

        for row in self.worksheet.iter_rows(min_row=2, values_only=False):  
            if row[0].value == book_to_edit:
                print("Book found! Enter new details:")
                row[0].value = input("Enter new book title: ").strip()
                row[1].value = input("Enter new author name: ").strip()
                row[2].value = input("Enter new publication year: ").strip()
                found = True
                break

        if found:
            self.save_workbook()
            print(f"Book '{book_to_edit}' has been updated successfully!")
        else:
            print("Book not found.")

    def display_menu(self):
        """Display user choices."""
        while True:
            print("\nLibrary Management Menu:")
            print("1. Add New Book")
            print("2. Edit Existing Book")
            print("3. View All Books")  
            print("4. Save and Exit")

            choice = input("Enter your choice (1-6): ")

            if choice == '1':
                self.add_book()
            elif choice == '2':
                self.edit_book()
            elif choice == '3':
                self.view_books()
            elif choice == '4':
                print("Saving changes and exiting...")
                self.save_workbook()
                break
            else:
                print("Invalid choice. Please try again.")

def main():
    manager = LibraryManager()
    manager.display_menu()

if __name__ == "__main__":
    main()
